from openpyxl import load_workbook

# Load the Excel workbook
workbook = load_workbook('your_excel_file.xlsx')

# Select the worksheet
worksheet = workbook.active

# Specify the range of cells you want to merge
start_row = 1
end_row = 3
start_column = 1
end_column = 3

# Get the values from the cells within the range
merged_value = ''
for row in range(start_row, end_row + 1):
    for column in range(start_column, end_column + 1):
        cell_value = worksheet.cell(row=row, column=column).value
        if cell_value:
            merged_value += str(cell_value) + ' '

# Set the merged value to the upper-left cell in the range
worksheet.cell(row=start_row, column=start_column, value=merged_value.strip())

# Merge the cells
worksheet.merge_cells(start_row=start_row, end_row=end_row, start_column=start_column, end_column=end_column)

# Save the workbook
workbook.save('your_modified_excel_file.xlsx')